#!/usr/bin/env python3
"""
Aonami Echo — Excel export of all platform data.
Produces AonamiEcho_Data_Export.xlsx with 8 sheets mirroring the dashboard.
"""
import sqlite3, json, math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import os

ROOT = os.path.dirname(os.path.abspath(__file__))
DB   = os.path.join(ROOT, "usfd.db")
OUT  = os.path.join(ROOT, "AonamiEcho_Data_Export.xlsx")

# ── colours ──────────────────────────────────────────────────────────
NAVY        = "0B1F3A"
NAVY_MID    = "132D52"
SAFFRON     = "FF7A1A"
RED         = "C8102E"
AMBER       = "E5A100"
GREEN       = "1F7A4D"
SLATE       = "506070"
LIGHT_RED   = "FFF0F2"
LIGHT_AMBER = "FFF8E6"
LIGHT_GREEN = "F0FAF5"
LIGHT_BLUE  = "EEF3FA"
BG_GRAY     = "F0F2F5"
WHITE       = "FFFFFF"
BORDER_COL  = "D8DCE3"

def rows(sql, params=()):
    con = sqlite3.connect(DB); con.row_factory = sqlite3.Row
    try:
        return [dict(r) for r in con.execute(sql, params).fetchall()]
    finally:
        con.close()

def fin(x):
    if x is None: return None
    try:
        v = float(x)
        return None if (math.isnan(v) or math.isinf(v) or v > 1e8) else round(v, 2)
    except Exception:
        return None

# ── style helpers ─────────────────────────────────────────────────────
def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hfont(hex_color=None, bold=False, size=10, name="Calibri"):
    kwargs = dict(name=name, size=size, bold=bold)
    if hex_color: kwargs["color"] = hex_color
    return Font(**kwargs)

def thin_border():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, col_names, row=1, fill_hex=NAVY, font_hex=WHITE, start_col=1):
    """Write a styled header row."""
    for ci, name in enumerate(col_names, start=start_col):
        c = ws.cell(row=row, column=ci, value=name)
        c.fill = hfill(fill_hex)
        c.font = hfont(font_hex, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 28

def style_data_cell(cell, bold=False, align="left", num_fmt=None, color=None, bg=None):
    cell.font = Font(name="Calibri", size=9, bold=bold,
                     color=color if color else "000000")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if num_fmt: cell.number_format = num_fmt
    if bg: cell.fill = hfill(bg)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws, freeze="A2", auto_filter=None):
    ws.freeze_panes = freeze
    if auto_filter:
        ws.auto_filter.ref = auto_filter

def triage_color(level):
    m = {"IMMEDIATE": (RED, LIGHT_RED),
         "URGENT": (AMBER, LIGHT_AMBER),
         "MONITOR": ("3e9dd6", "F0F8FF"),
         "ROUTINE": (GREEN, LIGHT_GREEN)}
    return m.get(level, (SLATE, WHITE))

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / EXECUTIVE SUMMARY
# ═══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 22

# Title block
ws.merge_cells("B2:F2")
c = ws["B2"]
c.value = "AONAMI ECHO  ·  USFD Rail & Weld Defect Triage"
c.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
c.fill = hfill(NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 36

ws.merge_cells("B3:F3")
c = ws["B3"]
c.value = "Indian Railways — Run-on-Run Defect Growth Propagation Analysis"
c.font = Font(name="Calibri", size=11, color="B0C0D0")
c.fill = hfill(NAVY_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 22

# Load stats
st_rows = rows("SELECT triage_level, COUNT(*) n FROM triage_queue GROUP BY triage_level")
st = {r["triage_level"]: r["n"] for r in st_rows}
near_imr = rows("SELECT COUNT(*) n FROM triage_queue WHERE latest_amp>=80")[0]["n"]
clusters_n = rows("SELECT COUNT(*) n FROM clusters")[0]["n"]
sla_br = rows("SELECT COUNT(*) n FROM triage_queue WHERE sla_breach=1")[0]["n"]
total = rows("SELECT COUNT(*) n FROM triage_queue")[0]["n"]
railways_n = rows("SELECT COUNT(DISTINCT railway) n FROM triage_queue")[0]["n"]
insp_n = rows("SELECT COUNT(*) n FROM inspections")[0]["n"]

kpis = [
    ("Total defects tracked",   total,                 NAVY,   WHITE),
    ("Inspection readings",     insp_n,                SLATE,  WHITE),
    ("Railways covered",        railways_n,            SLATE,  WHITE),
    ("IMMEDIATE (act now)",     st.get("IMMEDIATE",0), RED,    WHITE),
    ("URGENT (≤2yr to IMR)",    st.get("URGENT",0),   AMBER,  WHITE),
    ("MONITOR (≤5yr to IMR)",   st.get("MONITOR",0),  "3e9dd6", WHITE),
    ("ROUTINE (stable)",        st.get("ROUTINE",0),   GREEN,  WHITE),
    ("Near-IMR ≥80%FSH",       near_imr,              RED,    WHITE),
    ("4-metre clusters",        clusters_n,            SAFFRON, WHITE),
    ("SLA breaches",            sla_br,                RED,    WHITE),
]

for i, (label, val, bg, fg) in enumerate(kpis):
    row = 5 + i
    ws.row_dimensions[row].height = 24
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = Font(name="Calibri", size=10, color="506070")
    lc.alignment = Alignment(vertical="center")
    vc = ws.cell(row=row, column=3, value=val)
    vc.font = Font(name="Calibri", size=13, bold=True, color=bg)
    vc.alignment = Alignment(horizontal="center", vertical="center")

# Mini bar for triage breakdown
ws.row_dimensions[16].height = 20
ws["B16"].value = "TRIAGE BREAKDOWN"
ws["B16"].font = Font(name="Calibri", size=9, bold=True, color=SLATE)

triage_order = ["IMMEDIATE","URGENT","MONITOR","ROUTINE"]
triage_colors_fill = [RED, AMBER, "3e9dd6", GREEN]
for ci, (lvl, col) in enumerate(zip(triage_order, triage_colors_fill), start=2):
    n = st.get(lvl, 0)
    c = ws.cell(row=17, column=ci, value=f"{lvl}\n{n}")
    c.fill = hfill(col)
    c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws.row_dimensions[17].height = 32
    ws.column_dimensions[get_column_letter(ci)].width = 22

# Severity distribution
ws.row_dimensions[20].height = 20
ws["B20"].value = "AMPLITUDE BANDS (RDSO Annexure II-A)"
ws["B20"].font = Font(name="Calibri", size=9, bold=True, color=SLATE)

amp_bands = rows("""
    SELECT
      SUM(CASE WHEN latest_amp < 30 THEN 1 ELSE 0 END) lt30,
      SUM(CASE WHEN latest_amp>=30 AND latest_amp<60 THEN 1 ELSE 0 END) b30_60,
      SUM(CASE WHEN latest_amp>=60 AND latest_amp<80 THEN 1 ELSE 0 END) b60_80,
      SUM(CASE WHEN latest_amp>=80 THEN 1 ELSE 0 END) gte80
    FROM triage_queue
""")[0]
band_data = [
    ("<30%FSH — Low",    amp_bands["lt30"],   GREEN),
    ("30–60%FSH — Monitor", amp_bands["b30_60"], "3e9dd6"),
    ("60–80%FSH — Urgent",  amp_bands["b60_80"], AMBER),
    ("≥80%FSH — IMR",   amp_bands["gte80"],   RED),
]
for ci, (label, val, col) in enumerate(band_data, start=2):
    c = ws.cell(row=21, column=ci, value=f"{label}\n{val:,}")
    c.fill = hfill(col)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    ws.row_dimensions[21].height = 32

ws["B24"].value = (
    "Generated by Aonami Echo  ·  Challenge IC0000000177  ·  "
    "Defect Growth Propagation Analysis in Rails and Welds, Indian Railways"
)
ws["B24"].font = Font(name="Calibri", size=8, color="A0A0A0", italic=True)
ws.merge_cells("B24:F24")

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2 — PRIORITY QUEUE (full triage queue)
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Priority Queue")
ws2.sheet_view.showGridLines = False

cols = [
    "Rank","Triage Level","Railway","Section","Line",
    "KM","Meter","Rail","Asset",
    "Latest %FSH","Severity Band",
    "Yrs→80%FSH","GMT→80%FSH",
    "Cluster ID","Para 6.3.2","SLA Breach","Days Since Insp",
    "Readings","Reasons"
]
header_row(ws2, cols, row=1)
set_col_widths(ws2, [6,13,22,20,6, 8,8,6,7, 11,14, 10,10, 9,9,9,13, 9,42])
freeze_and_filter(ws2, "A2", f"A1:{get_column_letter(len(cols))}1")

q = rows("""SELECT tq.*, p.yrs_to_80, p.gmt_to_80
            FROM triage_queue tq
            LEFT JOIN predictions p ON tq.defect_key=p.defect_key
            ORDER BY tq.triage_rank""")

for ri, r in enumerate(q, start=2):
    lvl = r.get("triage_level","ROUTINE")
    fc, bg = triage_color(lvl)
    amp = r.get("latest_amp") or 0
    amp_bg = LIGHT_RED if amp >= 80 else (LIGHT_AMBER if amp >= 60 else WHITE)

    vals = [
        r.get("triage_rank"),
        lvl,
        r.get("railway",""),
        r.get("section",""),
        r.get("line",""),
        r.get("km"),
        r.get("meter"),
        r.get("rail",""),
        r.get("asset_type",""),
        fin(amp),
        r.get("severity_band",""),
        fin(r.get("yrs_to_80")),
        fin(r.get("gmt_to_80")),
        r.get("cluster_id") if (r.get("cluster_id") or -1) > 0 else "",
        "YES" if r.get("para632_escalate") else "No",
        "YES" if r.get("sla_breach") else "No",
        r.get("days_since_insp"),
        r.get("n_readings"),
        (r.get("reasons","") or "").replace("|"," · ")
    ]
    for ci, v in enumerate(vals, start=1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        c.font = Font(name="Calibri", size=9)

        if ci == 2:  # triage level
            c.fill = hfill(bg)
            c.font = Font(name="Calibri", size=9, bold=True, color=fc)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 10:  # amplitude
            c.fill = hfill(amp_bg)
            afc = RED if amp >= 80 else (AMBER if amp >= 60 else (SLATE if amp >= 30 else GREEN))
            c.font = Font(name="Calibri", size=9, bold=amp>=60, color=afc)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci in (15, 16) and v == "YES":
            c.fill = hfill(LIGHT_RED)
            c.font = Font(name="Calibri", size=9, bold=True, color=RED)
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[ri].height = 16

# ── Conditional formatting: amplitude colour scale ──
ws2.conditional_formatting.add(
    f"J2:J{len(q)+1}",
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="63BE7B",
        mid_type="num",   mid_value=60,    mid_color="FFEB84",
        end_type="num",   end_value=100,   end_color="F8696B"
    )
)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3 — IMMEDIATE ACTIONS (566 defects)
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("IMMEDIATE Actions")
ws3.sheet_view.showGridLines = False

cols3 = [
    "Rank","Railway","Section","KM","Meter","Rail","Asset",
    "Latest %FSH","Para 6.3.2 Cluster","SLA Breach","Days Since Insp",
    "Reasons","Recommended Action"
]
header_row(ws3, cols3, row=1, fill_hex=RED)
set_col_widths(ws3, [6,22,18,8,8,6,7, 11,14,9,14, 40,38])
freeze_and_filter(ws3, "A2", f"A1:{get_column_letter(len(cols3))}1")

ws3.row_dimensions[1].height = 28

imm = rows("""SELECT * FROM triage_queue WHERE triage_level='IMMEDIATE' ORDER BY triage_rank""")

action_map = {
    True: "CLUSTER — Single rail-piece replacement (Para 6.3.2). Raise one work order for entire cluster span. Place joggled fish plate immediately.",
    False: "IMR-LEVEL — Remove from track within 3 days (RDSO Para 6.2.3). Joggled fish plate in 24hrs. File Annexure III."
}

for ri, r in enumerate(imm, start=2):
    amp = r.get("latest_amp") or 0
    is_cluster = bool(r.get("para632_escalate"))
    is_breach = bool(r.get("sla_breach"))
    vals = [
        r.get("triage_rank"),
        r.get("railway",""),
        r.get("section",""),
        r.get("km"),
        r.get("meter"),
        r.get("rail",""),
        r.get("asset_type",""),
        fin(amp),
        "YES — CLUSTER" if is_cluster else "No",
        "OVERDUE" if is_breach else "Pending",
        r.get("days_since_insp"),
        (r.get("reasons","") or "").replace("|"," · "),
        action_map[is_cluster]
    ]
    for ci, v in enumerate(vals, start=1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(ci in (12,13)),
                                horizontal="right" if isinstance(v,(int,float)) else "left")

        if ci == 8:  # amplitude
            c.fill = hfill(LIGHT_RED)
            c.font = Font(name="Calibri", size=9, bold=True, color=RED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 9 and is_cluster:
            c.fill = hfill("FFF0F2")
            c.font = Font(name="Calibri", size=9, bold=True, color=RED)
        elif ci == 10:
            c.fill = hfill(LIGHT_RED if is_breach else LIGHT_AMBER)
            c.font = Font(name="Calibri", size=9, bold=True,
                          color=RED if is_breach else AMBER)
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws3.row_dimensions[ri].height = 28

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4 — 4-METRE CLUSTERS (Para 6.3.2)
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Clusters (Para 6.3.2)")
ws4.sheet_view.showGridLines = False

cols4 = [
    "Cluster ID","Railway","Section","Line",
    "Chainage Start (m)","Chainage End (m)","Span (m)",
    "No. Defects","Max Amplitude %FSH",
    "RDSO Rule","Replacement Action"
]
header_row(ws4, cols4, row=1, fill_hex=SAFFRON)
set_col_widths(ws4, [11,22,18,6, 16,14,9, 11,16, 18,46])
freeze_and_filter(ws4, "A2", f"A1:{get_column_letter(len(cols4))}1")

cl = rows("SELECT * FROM clusters ORDER BY n_defects DESC, max_amp DESC")
for ri, r in enumerate(cl, start=2):
    span = r.get("span_m") or 0
    n_def = r.get("n_defects") or 0
    max_amp = r.get("max_amp") or 0
    replace_len = round(span + 1.0, 1)
    vals = [
        f"CL-{r['cluster_id']:04d}",
        r.get("railway",""),
        r.get("section",""),
        r.get("line",""),
        fin(r.get("chainage_start_m")),
        fin(r.get("chainage_end_m")),
        fin(span),
        n_def,
        fin(max_amp),
        "Para 6.3.2 — 4-metre proximity",
        f"Single replacement ≥{replace_len}m covering all {n_def} defects. One Annexure III."
    ]
    amp_bg = LIGHT_RED if max_amp >= 80 else (LIGHT_AMBER if max_amp >= 60 else WHITE)
    for ci, v in enumerate(vals, start=1):
        c = ws4.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(ci==11),
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        if ci == 9:
            c.fill = hfill(amp_bg)
            c.font = Font(name="Calibri", size=9, bold=max_amp>=60, color=RED if max_amp>=80 else AMBER if max_amp>=60 else "000000")
    ws4.row_dimensions[ri].height = 28

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5 — RAILWAY ROLLUP
# ═══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Railway Rollup")
ws5.sheet_view.showGridLines = False

cols5 = [
    "Railway","Sections","Total Defects",
    "IMMEDIATE","URGENT","MONITOR","ROUTINE",
    "Near-IMR (≥80%)","60–80% (High)","30–60% (Watch)","<30% (Low)",
    "SLA Breaches","Avg %FSH","Max %FSH"
]
header_row(ws5, cols5, row=1, fill_hex=NAVY_MID)
set_col_widths(ws5, [26,10,14, 12,10,10,10, 15,13,14,10, 13,11,10])
freeze_and_filter(ws5, "A2")

zn = rows("""
    SELECT railway,
      COUNT(DISTINCT section) sections,
      COUNT(*) total,
      SUM(CASE WHEN triage_level='IMMEDIATE' THEN 1 ELSE 0 END) imm,
      SUM(CASE WHEN triage_level='URGENT'    THEN 1 ELSE 0 END) urg,
      SUM(CASE WHEN triage_level='MONITOR'   THEN 1 ELSE 0 END) mon,
      SUM(CASE WHEN triage_level='ROUTINE'   THEN 1 ELSE 0 END) rou,
      SUM(CASE WHEN latest_amp>=80 THEN 1 ELSE 0 END) near_imr,
      SUM(CASE WHEN latest_amp>=60 AND latest_amp<80 THEN 1 ELSE 0 END) hi,
      SUM(CASE WHEN latest_amp>=30 AND latest_amp<60 THEN 1 ELSE 0 END) watch,
      SUM(CASE WHEN latest_amp<30 THEN 1 ELSE 0 END) low,
      SUM(sla_breach) sla_br,
      ROUND(AVG(latest_amp),1) avg_amp,
      ROUND(MAX(latest_amp),1) max_amp
    FROM triage_queue GROUP BY railway ORDER BY near_imr DESC, imm DESC
""")

for ri, r in enumerate(zn, start=2):
    near_imr_val = r.get("near_imr", 0)
    imm_val = r.get("imm", 0)
    avg = r.get("avg_amp") or 0
    row_bg = LIGHT_RED if near_imr_val > 0 else (LIGHT_AMBER if imm_val > 0 else WHITE)
    vals = [
        r.get("railway",""),
        r.get("sections"),
        r.get("total"),
        imm_val,
        r.get("urg"),
        r.get("mon"),
        r.get("rou"),
        near_imr_val,
        r.get("hi"),
        r.get("watch"),
        r.get("low"),
        r.get("sla_br"),
        fin(avg),
        fin(r.get("max_amp"))
    ]
    for ci, v in enumerate(vals, start=1):
        c = ws5.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        if ci == 4 and v:   # IMMEDIATE
            c.fill = hfill(LIGHT_RED); c.font = Font(name="Calibri", size=10, bold=bool(v), color=RED)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 5 and v: # URGENT
            c.fill = hfill(LIGHT_AMBER); c.font = Font(name="Calibri", size=10, bold=bool(v), color=AMBER)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 8:       # near-IMR
            c.fill = hfill(LIGHT_RED if v else WHITE)
            c.font = Font(name="Calibri", size=10, bold=bool(v), color=RED if v else "000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 13:      # avg amp
            c.fill = hfill(LIGHT_RED if avg>=60 else LIGHT_AMBER if avg>=40 else WHITE)
            c.font = Font(name="Calibri", size=10, bold=avg>=60, color=RED if avg>=60 else AMBER if avg>=40 else "000000")
        elif ci == 12 and v: # SLA
            c.fill = hfill(LIGHT_RED); c.font = Font(name="Calibri", size=10, bold=True, color=RED)
    ws5.row_dimensions[ri].height = 18

# Totals row
ws5.row_dimensions[len(zn)+2].height = 20
totals = [
    "FLEET TOTAL","","",
    sum(r.get("imm",0) for r in zn),
    sum(r.get("urg",0) for r in zn),
    sum(r.get("mon",0) for r in zn),
    sum(r.get("rou",0) for r in zn),
    sum(r.get("near_imr",0) for r in zn),
    sum(r.get("hi",0) for r in zn),
    sum(r.get("watch",0) for r in zn),
    sum(r.get("low",0) for r in zn),
    sum(r.get("sla_br",0) for r in zn),
    None, None
]
totals[2] = sum(r.get("total",0) for r in zn)
tr = len(zn) + 2
for ci, v in enumerate(totals, start=1):
    c = ws5.cell(row=tr, column=ci, value=v)
    c.fill = hfill(NAVY); c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    c.border = thin_border()
    c.alignment = Alignment(vertical="center", horizontal="right" if isinstance(v,(int,float)) else "left")

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6 — RUL PROJECTIONS (all defects with model outputs)
# ═══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("RUL Projections")
ws6.sheet_view.showGridLines = False

cols6 = [
    "Defect Key","Railway","Section","KM","Meter","Rail","Asset","Family",
    "Latest %FSH","Own Slope (%/GMT)","Readings",
    "Yrs→30%FSH","GMT→30%FSH",
    "Yrs→60%FSH","GMT→60%FSH",
    "Yrs→80%FSH (IMR)","GMT→80%FSH (IMR)",
    "Risk Score","Triage Level"
]
header_row(ws6, cols6, row=1, fill_hex=SLATE)
set_col_widths(ws6, [38,22,18,8,8,6,8,14, 11,16,9, 11,11, 11,11, 14,12, 10,13])
freeze_and_filter(ws6, "A2", f"A1:{get_column_letter(len(cols6))}1")

preds = rows("""
    SELECT p.*, t.triage_level, t.severity_band
    FROM predictions p
    LEFT JOIN triage_queue t ON p.defect_key=t.defect_key
    ORDER BY p.risk_rank
""")

for ri, r in enumerate(preds, start=2):
    amp = r.get("latest_amp") or 0
    lvl = r.get("triage_level","ROUTINE")
    y80 = fin(r.get("yrs_to_80"))
    rs_raw = r.get("risk_score") or 0
    rs = min(1.0, rs_raw / 1e6) if rs_raw > 1e5 else round(min(1.0, rs_raw), 3)
    slope = fin(r.get("own_slope_per_gmt"))

    vals = [
        r.get("defect_key",""),
        r.get("railway",""),
        r.get("section",""),
        r.get("km"),
        r.get("meter"),
        r.get("rail",""),
        r.get("asset_type",""),
        r.get("family",""),
        fin(amp),
        slope,
        r.get("n_readings"),
        fin(r.get("yrs_to_30")),
        fin(r.get("gmt_to_30")),
        fin(r.get("yrs_to_60")),
        fin(r.get("gmt_to_60")),
        y80,
        fin(r.get("gmt_to_80")),
        rs,
        lvl
    ]
    fc, bg = triage_color(lvl)
    amp_bg = LIGHT_RED if amp >= 80 else (LIGHT_AMBER if amp >= 60 else WHITE)

    for ci, v in enumerate(vals, start=1):
        c = ws6.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=8)
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        if ci == 9:
            c.fill = hfill(amp_bg)
        elif ci == 16 and y80 is not None:
            rul_bg = LIGHT_RED if y80 < 1 else (LIGHT_AMBER if y80 < 3 else WHITE)
            rul_fc = RED if y80 < 1 else (AMBER if y80 < 3 else GREEN)
            c.fill = hfill(rul_bg)
            c.font = Font(name="Calibri", size=8, bold=y80<2, color=rul_fc)
        elif ci == 19:
            c.fill = hfill(bg)
            c.font = Font(name="Calibri", size=8, bold=True, color=fc)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[ri].height = 14

ws6.conditional_formatting.add(
    f"I2:I{len(preds)+1}",
    ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",
                   mid_type="num",  mid_value=60,   mid_color="FFEB84",
                   end_type="num",  end_value=100,  end_color="F8696B")
)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 7 — PARIS–ERDOGAN FAMILY MODELS
# ═══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Growth Model (Paris-Erdogan)")
ws7.sheet_view.showGridLines = False

cols7 = [
    "Family","Asset Type","Probe Angle",
    "C (coefficient)","m (exponent)",
    "Positive Growth Segments","Defects in Family",
    "Avg Amplitude %FSH","Near-IMR (≥80%)",
    "Growth Rate at 30%FSH","Growth Rate at 60%FSH","Growth Rate at 80%FSH",
    "Interpretation"
]
header_row(ws7, cols7, row=1, fill_hex="132D52")
set_col_widths(ws7, [16,12,12, 16,14, 22,17, 18,15, 18,18,18, 52])

fams = rows("""
    SELECT fp.family, fp.C, fp.m, fp.n_pos_segments,
           COUNT(p.defect_key) n_defects,
           ROUND(AVG(p.latest_amp),1) avg_amp,
           SUM(CASE WHEN p.latest_amp>=80 THEN 1 ELSE 0 END) near_imr
    FROM family_powerlaw fp
    LEFT JOIN predictions p ON p.family=fp.family
    GROUP BY fp.family
    ORDER BY avg_amp DESC
""")

def interpret_family(fam, C, m, avg_amp):
    asset = "Rail" if fam.startswith("rail") else "Weld" if fam.startswith("weld") else "Global"
    probe = fam.split(":")[-1].upper() if ":" in fam else "all"
    if m > 1.5:
        growth = "strongly accelerating — high amplitude defects grow much faster"
    elif m > 1.0:
        growth = "moderately accelerating — growth rate increases with amplitude"
    elif m > 0.6:
        growth = "mildly accelerating — nearly linear growth with amplitude"
    else:
        growth = "sub-linear — growth rate decreases at high amplitudes (probe artefact likely)"
    risk = "HIGH risk" if (avg_amp or 0) >= 40 else "Moderate risk" if (avg_amp or 0) >= 30 else "Lower risk"
    return f"{risk}. {asset} defects detected by {probe} probe. Growth {growth}. da/dN = {C:.2e} · a^{m:.2f}."

for ri, r in enumerate(fams, start=2):
    C = r.get("C") or 0
    m = r.get("m") or 0
    avg = r.get("avg_amp") or 0
    fam = r.get("family","")
    parts = fam.split(":") if ":" in fam else [fam, "all"]
    asset = parts[0].capitalize()
    probe = parts[1].upper() if len(parts)>1 else "all"
    rate30 = round(C * (30**m), 4) if C and m else None
    rate60 = round(C * (60**m), 4) if C and m else None
    rate80 = round(C * (80**m), 4) if C and m else None

    vals = [
        fam, asset, probe,
        C, m,
        r.get("n_pos_segments"), r.get("n_defects"),
        avg, r.get("near_imr"),
        rate30, rate60, rate80,
        interpret_family(fam, C, m, avg)
    ]
    for ci, v in enumerate(vals, start=1):
        c = ws7.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=9)
        c.alignment = Alignment(vertical="center", wrap_text=(ci==13),
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        if ci in (4,5):
            c.number_format = "0.000E+00" if ci==4 else "0.0000"
            c.font = Font(name="Calibri", size=9, bold=True, color="1C3F6E")
        elif ci == 8:
            c.fill = hfill(LIGHT_RED if avg>=50 else LIGHT_AMBER if avg>=35 else WHITE)
            c.font = Font(name="Calibri", size=9, bold=avg>=50, color=RED if avg>=50 else AMBER if avg>=40 else "000000")
        elif ci in (10,11,12) and v:
            c.number_format = "0.0000"
            high_rate = v > 2
            c.fill = hfill(LIGHT_RED if high_rate else WHITE)
            c.font = Font(name="Calibri", size=9, bold=high_rate, color=RED if high_rate else "000000")
    ws7.row_dimensions[ri].height = 48

# ═══════════════════════════════════════════════════════════════════════
# SHEET 8 — MODEL SCORECARD (HistGBM + validation)
# ═══════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Model Scorecard")
ws8.sheet_view.showGridLines = False
ws8.column_dimensions["A"].width = 2
ws8.column_dimensions["B"].width = 40
ws8.column_dimensions["C"].width = 30
ws8.column_dimensions["D"].width = 30

meta = json.load(open(os.path.join(ROOT, "model_meta.json")))

def ws8_row(row, label, value, label_bg=None, val_bg=None, bold_val=False):
    lc = ws8.cell(row=row, column=2, value=label)
    vc = ws8.cell(row=row, column=3, value=value)
    lc.font = Font(name="Calibri", size=10, color=SLATE)
    vc.font = Font(name="Calibri", size=10, bold=bold_val)
    lc.alignment = Alignment(vertical="center")
    vc.alignment = Alignment(vertical="center")
    lc.border = thin_border(); vc.border = thin_border()
    if label_bg: lc.fill = hfill(label_bg)
    if val_bg: vc.fill = hfill(val_bg)
    ws8.row_dimensions[row].height = 20

# Title
ws8.merge_cells("B2:D2")
t = ws8["B2"]
t.value = "HistGradientBoosting — Next-Reading Amplitude Predictor"
t.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
t.fill = hfill(NAVY); t.alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[2].height = 30

section_titles = {
    4: "TARGET & FEATURES", 8: "TRAINING DATA", 12: "5-FOLD GROUPKFOLD CV RESULTS",
    18: "BASELINE COMPARISON", 24: "INTERPRETATION"
}
for sr, title in section_titles.items():
    ws8.merge_cells(f"B{sr}:D{sr}")
    c = ws8.cell(row=sr, column=2, value=title)
    c.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
    c.fill = hfill(SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws8.row_dimensions[sr].height = 18

ws8_row(5,  "Target (output)",         meta["target"])
ws8_row(6,  "Numeric features",        ", ".join(meta["features_numeric"]))
ws8_row(7,  "Categorical features",    ", ".join(meta["features_categorical"]))
ws8_row(9,  "Training transitions",    f"{meta['n_training_transitions']:,}")
ws8_row(10, "Defects used",            f"{meta['n_defects']:,}")
ws8_row(11, "Validation strategy",     meta["cv"])

ws8_row(13, "HistGBM MAE",             f"{meta['model_mae']} %FSH",      val_bg=LIGHT_BLUE)
ws8_row(14, "HistGBM R²",              f"{meta['model_r2']}",             val_bg=LIGHT_BLUE)
ws8_row(15, "Persistence baseline MAE",f"{meta['persistence_mae']} %FSH", val_bg=BG_GRAY)
ws8_row(16, "Persistence baseline R²", f"{meta['persistence_r2']}",       val_bg=BG_GRAY)
ws8_row(17, "HistGBM beats persistence on MAE?",
        "NO — persistence wins on MAE" if not meta["beats_persistence_on_mae"] else "YES",
        val_bg=LIGHT_RED if not meta["beats_persistence_on_mae"] else LIGHT_GREEN,
        bold_val=True)

ws8_row(19, "Holdout test (1,008 defects)",
        "Persistence MAE 4.21 %FSH / ML MAE 4.58 %FSH", val_bg=LIGHT_AMBER)
ws8_row(20, "Holdout split method",    "hashlib.md5(defect_key) — deterministic 70/30")
ws8_row(21, "Linear regression MAE",  "6.27 %FSH")
ws8_row(22, "Physics power-law MAE",  "8.80 %FSH")
ws8_row(23, "Best by MAE",            "Persistence (assume unchanged)", bold_val=True)

interpret = (
    "The ML model achieves higher R² (0.687 vs 0.535) but does NOT beat persistence on MAE. "
    "This is expected: with ≤3 readings per defect, ~54% of defects are flat (median delta=0), "
    "so persistence is correct for most of them. The model's value is in RANKING "
    "(which defects are most likely to grow) and in the RUL projection, not in precise "
    "next-amplitude point forecasts. Treat predicted IMR dates as ±1yr estimates, not precise dates."
)
ws8.merge_cells("B25:D28")
c = ws8["B25"]
c.value = interpret
c.font = Font(name="Calibri", size=10, color="506070", italic=True)
c.fill = hfill("FFFBF0")
c.alignment = Alignment(vertical="top", wrap_text=True)
c.border = thin_border()
ws8.row_dimensions[25].height = 18
for r in range(26, 29):
    ws8.row_dimensions[r].height = 18

# ═══════════════════════════════════════════════════════════════════════
# SHEET 9 — INSPECTION HISTORY (per-defect readings)
# ═══════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Inspection History")
ws9.sheet_view.showGridLines = False

cols9 = [
    "Defect Key","Railway","Section","Line","KM","Meter","Rail","Asset",
    "Round","Inspection Date","Classification","Probe","Amplitude %FSH","GMT at Round"
]
header_row(ws9, cols9, row=1, fill_hex=NAVY_MID)
set_col_widths(ws9, [38,22,18,6,7,8,6,8, 11,16,16,12,14,14])
freeze_and_filter(ws9, "A2", f"A1:{get_column_letter(len(cols9))}1")

hist = rows("""
    SELECT i.defect_key, i.railway, i.section, i.line, i.km, i.meter, i.rail, i.asset_type,
           i.round_label, i.inspection_date, i.classification, i.probe,
           i.amplitude_pct, i.gmt_at_round
    FROM inspections i
    ORDER BY i.railway, i.section, i.km, i.meter, i.insp_dt
""")

for ri, r in enumerate(hist, start=2):
    amp = r.get("amplitude_pct")
    amp_bg = LIGHT_RED if (amp or 0) >= 80 else (LIGHT_AMBER if (amp or 0) >= 60 else WHITE)
    vals = [
        r.get("defect_key",""), r.get("railway",""), r.get("section",""), r.get("line",""),
        r.get("km"), r.get("meter"), r.get("rail",""), r.get("asset_type",""),
        r.get("round_label",""), r.get("inspection_date",""),
        r.get("classification",""), r.get("probe",""),
        fin(amp), fin(r.get("gmt_at_round"))
    ]
    for ci, v in enumerate(vals, start=1):
        c = ws9.cell(row=ri, column=ci, value=v)
        c.border = thin_border()
        c.font = Font(name="Calibri", size=8)
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if isinstance(v,(int,float)) else "left")
        if ci == 13:
            c.fill = hfill(amp_bg)
            if amp and amp >= 60:
                c.font = Font(name="Calibri", size=8, bold=True, color=RED if amp>=80 else AMBER)
    ws9.row_dimensions[ri].height = 13

ws9.conditional_formatting.add(
    f"M2:M{len(hist)+1}",
    ColorScaleRule(start_type="num", start_value=0,  start_color="63BE7B",
                   mid_type="num",  mid_value=60,    mid_color="FFEB84",
                   end_type="num",  end_value=100,   end_color="F8696B")
)

# ═══════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════
wb.save(OUT)
print(f"✓  Saved: {OUT}")
sizes = []
for sh in wb.sheetnames:
    ws_tmp = wb[sh]
    rows_count = ws_tmp.max_row - 1  # exclude header
    print(f"   {sh:<32} {max(0,rows_count):>6,} data rows")
